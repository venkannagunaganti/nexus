import openpyxl
import csv
import xlrd
import os


# =============================================================================
# Function name: read_excel_and_write_to_file
# Author: <Methku Vikas>
# Description: Reads data from an Excel file (.xlsx or .xls) at the specified 'excel_file_path'
#              and writes it to either a CSV file or a TXT file based on the 'output_file_path' extension.
# Input Parameters:
#   excel_file_path (str) - The path to the Excel file to read data from.
# Output Parameters: output_file_path (str) - The path to the output file where data will be written. The file format
#                             will be determined based on the file extension (.csv or .txt).
# How to invoke?: CALL THIS Method ----- read_excel_and_write_to_file(excel_file_path, output_file_path)
# Date created: 02/08/2023
# Date last modified & Changes done: 02/08/2023
# =============================================================================

def read_excel_and_write_to_file(excel_file_path, output_file_path):
    # checking the file existing or not
    if not os.path.exists(excel_file_path):
        print("Excel file does not exists.")
    else:

        file_extension = excel_file_path.split('.')[-1].lower()

        if file_extension == 'xlsx':
            # Read .xlsx file using openpyxl
            wb = openpyxl.load_workbook(excel_file_path)
            sheet = wb.active
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
        elif file_extension == 'xls':
            # Read .xls file using xlrd
            wb = xlrd.open_workbook(excel_file_path)
            sheet = wb.sheet_by_index(0)
            data = [sheet.row_values(row_num) for row_num in range(sheet.nrows)]
        else:
            raise ValueError("Unsupported file format. Please use .xlsx or .xls files.")

        # Check the output file format based on the output_file_path extension
        output_extension = output_file_path.split('.')[-1].lower()

        if output_extension == 'csv':
            # Write the data to a CSV file
            with open(output_file_path, 'w', newline='') as csv_file:
                writer = csv.writer(csv_file)
                writer.writerows(data)
        elif output_extension == 'txt':
            # Write the data to a TXT file
            with open(output_file_path, 'w') as txt_file:
                for row in data:
                    txt_file.write('\t'.join(map(str, row)) + '\n')
        else:
            raise ValueError("Unsupported output format. Please use '.csv' or '.txt' as the output file extension.")


# ============================================================================= 
# Function name:  write_specific_columns_to_csv
# Author: <Methku vikas, Kondri Satyannarayana>
# Description: Reads data from an Excel file (.xlsx) at the specified 'excel_file_path' and
#              writes specific columns (specified by their indices) to a CSV file at the 'output_file_path'.
# Input Parameters: excel_file_path (str) - The path to the Excel file to read data from.
#                   columns_to_write (list) - A list of integers representing the indices of the columns to be written.
# Output Parameters: output_file_path (str) - The path to the output CSV file where selected columns will be written.
# How to invoke?: CALL THIS Method -- write_specific_columns_to_csv(excel_file_path, output_file_path, columns_to_write)
# Date created: 02/08/2023
# Date last modified & Changes done: 02/08/2023
# =============================================================================

def write_specific_columns_to_csv(excel_file_path, output_file_path, columns_to_write):
    # Read .xlsx file using openpyxl
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb.active

    # Extract the specified columns from the Excel file
    data = []
    for row in sheet.iter_rows(values_only=True):
        selected_data = [row[column - 1] for column in columns_to_write]
        data.append(selected_data)

    # Write the selected data to a CSV file
    with open(output_file_path, 'w', newline='') as csv_file:
        writer = csv.writer(csv_file)
        writer.writerows(data)


# =============================================================================
# Function name: add_columns_to_csv
# Author: <Methku Vikas, Kondri Satyannarayana>
# Description: Adds new columns with specified values to an existing CSV file at the 'csv_file_path'.
# Input Parameters:
#   csv_file_path (str) - The path to the existing CSV file.
#   columns_to_add (list) - A list of values representing the new columns to be added.
# Output Parameters: Same as input parameter
# How to invoke?: CALL THIS Method ----- add_columns_to_csv(csv_file_path, columns_to_add)
# Date created: 02/08/2023
# Date last modified & Changes done: 02/08/2023
# =============================================================================
def add_columns_to_csv(csv_file_path, columns_to_add):
    # Read the existing data from the CSV file
    existing_data = []
    with open(csv_file_path, 'r') as csv_file:
        reader = csv.reader(csv_file)
        for row in reader:
            existing_data.append(row)

    # Add new columns to the existing data
    for i in range(len(existing_data)):
        existing_data[i] += columns_to_add

    # Write the updated data to the CSV file
    with open(csv_file_path, 'w', newline='') as csv_file:
        writer = csv.writer(csv_file)
        writer.writerows(existing_data)


# =============================================================================
# Function name: add_rows_to_csv
# Author: <Methku Vikas, Kondri Satyannarayana>
# Description: Adds new rows with specified values to an existing CSV file at the 'csv_file_path'.
# Input Parameters: csv_file_path (str) - The path to the existing CSV file.
#   rows_to_add (list of lists) - A list of lists representing the new rows to be added.
# Output Parameters: same as input parameter
# How to invoke?: CALL THIS Method ----- add_rows_to_csv(csv_file_path, rows_to_add)
# Date created: 02/08/2023
# Date last modified & Changes done: 02/08/2023
# =============================================================================

def add_rows_to_csv(csv_file_path, rows_to_add):
    # Read the existing data from the CSV file
    existing_data = []
    with open(csv_file_path, 'r') as csv_file:
        reader = csv.reader(csv_file)
        for row in reader:
            existing_data.append(row)

    # Add new rows to the existing data
    existing_data += rows_to_add

    # Write the updated data to the CSV file
    with open(csv_file_path, 'w', newline='') as csv_file:
        writer = csv.writer(csv_file)
        writer.writerows(existing_data)


# =============================================================================
# Function name: write_csv_to_excel
# Author: <Methku Vikas, Kondri Satyannarayana>
# Description: Reads data from a CSV file at the 'csv_file_path' and writes it to an Excel file
#              at the 'output_excel_file_path' in the .xlsx format.
# Input Parameters: csv_file_path (str) - The path to the CSV file to read data from.
# Output Parameters: output_excel_file_path (str) - The path to the output Excel file where data will be written.
# How to invoke?: CALL THIS Method ----- write_csv_to_excel(csv_file_path, output_excel_file_path)
# Date created: 02/08/2023
# Date last modified & Changes done: 02/08/2023
# =============================================================================

def write_csv_to_excel(csv_file_path, output_excel_file_path):
    # Read data from the CSV file
    data = []
    with open(csv_file_path, 'r') as csv_file:
        reader = csv.reader(csv_file)
        for row in reader:
            data.append(row)

    # Write the data to the Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in data:
        ws.append(row)

    # Save the Excel file
    wb.save(output_excel_file_path)
